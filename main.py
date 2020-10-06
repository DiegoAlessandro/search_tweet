import requests
from bs4 import BeautifulSoup as BS
from selenium import webdriver
import platform
import openpyxl
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import urllib.parse
from PIL import Image
import datetime


def get_input_values():
    # 入力項目
    print('検索Twitterアカウント名：')
    twitter_user_name = input()

    input_start_date = None
    input_end_date = None
    try:
        print('検索開始日付(YYYYMMDD)：')
        input_start_date = input()
        input_start_date = datetime.datetime.strptime(input_start_date, "%Y%m%d").strftime("%Y-%m-%d")
    except:
        print('検索開始日付(YYYYMMDD)が不正です。')
        return twitter_user_name, None, None

    try:
        print('検索終了日付(YYYYMMDD)：')
        input_end_date = input()
        input_end_date = datetime.datetime.strptime(input_end_date, "%Y%m%d").strftime("%Y-%m-%d")
    except:
        print('検索終了日付(YYYYMMDD)が不正です。')
        return twitter_user_name, None, None

    return twitter_user_name, input_start_date, input_end_date


def start_chrome():
    options = webdriver.ChromeOptions()
    # options.add_argument('--headless')
    if platform.system() == 'Darwin':
        driver = webdriver.Chrome(options=options, executable_path='driver/mac/chromedriver')  # ローカルテスト用

    else:
        driver = webdriver.Edge(executable_path='driver/windows/msedgedriver.exe')

    return driver


def set_scroll(driver):
    # Get scroll height
    last_height = driver.execute_script("return document.body.scrollHeight")

    while True:
        # Scroll down to bottom
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

        # Calculate new scroll height and compare with last scroll height
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height


def download_image(image_url, timeout=10):
    response = requests.get(image_url, allow_redirects=False, timeout=timeout)
    if response.status_code != 200:
        e = Exception("HTTP status: " + str(response.status_code))
        raise e

    content_type = response.headers["content-type"]
    if 'image' not in content_type:
        e = Exception("Content-Type: " + content_type)
        raise e

    return response.content


# 画像を保存する
def save_image(filename, image):
    with open(filename, "wb") as fout:
        fout.write(image)
    img = Image.open(filename)
    if img.size[1] > img.size[0]:
        img.thumbnail((int(round((img.size[1] / 300) * img.size[0])), 300), Image.ANTIALIAS)
    else:
        img.thumbnail((300, int(round((img.size[0] / 300) * img.size[1]))), Image.ANTIALIAS)

    img.save(filename)


def generate_excel(export_file_name, exec_date, search_user_name, data):
    COL_WIDTH_RATIO = 1.2

    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]

    # 抽出日(B1)
    ws["B1"].value = "抽出日"
    ws["C1"].value = exec_date

    # アカウント(B2)
    ws["B2"].value = "アカウント"
    ws["C2"].value = search_user_name

    # set width column
    ws.column_dimensions['A'].width = 10 * COL_WIDTH_RATIO
    ws.column_dimensions['B'].width = 10 * COL_WIDTH_RATIO
    ws.column_dimensions['C'].width = 15 * COL_WIDTH_RATIO
    ws.column_dimensions['D'].width = 15 * COL_WIDTH_RATIO
    ws.column_dimensions['E'].width = 30 * COL_WIDTH_RATIO
    ws.column_dimensions['F'].width = 40 * COL_WIDTH_RATIO
    ws.column_dimensions['G'].width = 5 * COL_WIDTH_RATIO
    ws.column_dimensions['H'].width = 5 * COL_WIDTH_RATIO

    # ヘッダの記載
    for _ in data:
        row_num = 4
        ws.cell(row_num, 1, "ツイートURL")
        ws.cell(row_num, 2, "ツイート日")
        ws.cell(row_num, 3, "ツイート種別")
        ws.cell(row_num, 4, "リツイート元アカウント")
        ws.cell(row_num, 5, "投稿テキスト")
        ws.cell(row_num, 6, "投稿画像")
        ws.cell(row_num, 7, "リツイート数")
        ws.cell(row_num, 8, "いいね数")

    # ボディの記載
    for i, d in enumerate(data):
        row_num = 4 + i + 1
        ws.cell(row_num, 1, '=HYPERLINK("{}", "{}")'.format(d.get('tweet_url'), d.get('tweet_url')))
        ws.cell(row_num, 2, d.get('tweet_time'))
        ws.cell(row_num, 3, d.get('tweet_type'))
        ws.cell(row_num, 4, d.get('tweet_user_name'))
        ws.cell(row_num, 5, d.get('tweet_postText'))
        ws.cell(row_num, 6, '')
        ws.cell(row_num, 7, d.get('retweet_num'))
        ws.cell(row_num, 8, d.get('like_num'))
        ws.row_dimensions[row_num].height = 170
        ws['E' + str(row_num)].alignment = openpyxl.styles.Alignment(wrapText=True)

        # 画像ダウンロード
        for image_url in d.get('image_urls'):
            # ダウンロード
            qs = urllib.parse.urlparse(image_url).query
            format = urllib.parse.parse_qs(qs).get('format')[0]
            image_b = download_image(image_url)
            filename = "./images/{}.{}".format(i, format)
            save_image(filename, image_b)

            # ファイル保存
            # Excelにファイルを配置
            img = openpyxl.drawing.image.Image(filename)
            img.anchor = 'F' + str(row_num)
            ws.add_image(img)

            # ファイル削除
            # os.remove(filename)

        wb.save("{}.xlsx".format(export_file_name))


def do_scrape(driver):
    # 全ツイートを表示するまでスクロール
    # 最初のページのツイートを取得する
    wait = WebDriverWait(driver, 120)
    element = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@aria-label="タイムライン: タイムラインを検索"]')))
    soup = BS(driver.page_source, 'lxml')
    section = soup.find('div', attrs={'aria-label': "タイムライン: タイムラインを検索"})
    tweets_divs = section.div.contents
    for t in tweets_divs:
        if t.article is None:
            continue

        # ツイートURL
        tweet_url = 'https://twitter.com{}'.format(
            t.article.div.div.div.contents[1].contents[1].contents[0].div.div.div.contents[2].attrs.get('href'))
        # ツイートID
        tweet_id = tweet_url.split('/')[-1]

        # アカウント名
        user_name = \
            t.article.div.div.div.contents[1].contents[1].contents[0].contents[0].div.div.div.div.contents[1].text
        # ツイート日
        tweet_time = datetime.datetime.strptime(t.time.attrs.get('datetime'), '%Y-%m-%dT%H:%M:%S.%fZ').strftime(
            '%Y年%m月%d日')

        # ツイート種別
        tweet_type = ''
        if t.article.div.div.div.contents[1].contents[1].contents[1].contents[1].div is not None:
            retweet_div = t.article.div.div.div.contents[1].contents[1].contents[1].contents[1].div.div.div.div.span
            if retweet_div is not None:
                tweet_type = "リツイート" if retweet_div.text == '引用ツイート' else 'オーガニックツイート'
            else:
                tweet_type = "オーガニックツイート"
        else:
            tweet_type = "オーガニックツイート"

        # リツイート元アカウント
        tweet_user_name = \
            t.article.div.div.div.contents[1].contents[1].contents[0].contents[0].div.div.div.div.contents[0].text

        # 投稿テキスト
        tweet_postText = t.article.div.div.div.contents[1].contents[1].contents[1].contents[0].text

        # 投稿画像
        image_urls = []
        if t.article.div.div.div.contents[1].contents[1].contents[1].contents[1].div is not None:
            if t.article.div.div.div.contents[1].contents[1].contents[1].contents[
                1].div.div.div.div.div is not None:
                if len(t.article.div.div.div.contents[1].contents[1].contents[1].contents[
                           1].div.div.div.div.div.contents) > 0 and \
                        t.article.div.div.div.contents[1].contents[1].contents[1].contents[
                            1].div.div.div.div.div.contents[1].div is not None:
                    images_url_div = \
                        t.article.div.div.div.contents[1].contents[1].contents[1].contents[
                            1].div.div.div.div.div.contents[
                            1].div.contents
                    for i in images_url_div:
                        if i.name == 'img':
                            image_urls.append(i.attrs.get('src'))

        # リツイート数
        retweet_num = t.article.div.div.div.contents[1].contents[1].contents[1].contents[2].contents[1].text
        retweet_num = retweet_num if retweet_num != '' else 0

        # いいね数
        like_num = t.article.div.div.div.contents[1].contents[1].contents[1].contents[2].contents[2].text
        like_num = like_num if like_num != '' else 0

        yield {tweet_id: {"tweet_id": tweet_id, "user_name": user_name, "tweet_type": tweet_type,
                          "tweet_time": tweet_time,
                          "tweet_user_name": tweet_user_name, "tweet_postText": tweet_postText,
                          "image_urls": image_urls, "retweet_num": retweet_num, "like_num": like_num,
                          "tweet_url": tweet_url}}


def execute_search():
    search_result = {}
    step_value = driver.execute_script("return document.body.scrollHeight") * 0.9
    current_height = 0
    while True:
        # Get scroll height
        last_height = driver.execute_script("return document.body.scrollHeight")

        # スクレイピングの実行
        for tweet_rs in do_scrape(driver):
            search_result.update(tweet_rs)

        # Scroll down to bottom
        current_height += step_value
        driver.execute_script("window.scrollBy(0, {});".format(step_value))
        time.sleep(1)

        # Calculate new scroll height and compare with last scroll height
        max_scroll_height = driver.execute_script("return document.body.scrollHeight")
        if current_height >= last_height:
            break

        last_height = max_scroll_height

    no_duplicate = list(set(search_result.keys()))
    r = []
    [r.append(search_result[i]) for i in no_duplicate]
    return r


if __name__ == '__main__':
    # 入力
    user_name, start_data, end_data = get_input_values()
    if start_data is None or end_data is None:
        exit()

    # selenium生成
    search_url = 'https://twitter.com/search?q=from%3A{}%20since%3A{}%20until%3A{}%20include%3Aretweets%20include%3Anativeretweets%20-filter%3Areplies%20-filter%3Aretweets&src=typed_query&f=live'.format(
        user_name, start_data, end_data)
    print(search_url)
    driver = start_chrome()
    driver.get(search_url)

    # 検索
    search_result = execute_search()
    print('件数：{}'.format(len(search_result)))

    # Excel出力
    exec_datetime = datetime.datetime.now().strftime("%Y年%m月%d日")
    generate_excel(export_file_name='{}_{}'.format(user_name, exec_datetime), exec_date=exec_datetime,
                   search_user_name=user_name,
                   data=search_result)

    driver.close()
    driver.quit()
    print('success!!')
